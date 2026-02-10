"""Excel (XLSX) import and export for parts and jobs."""

from pathlib import Path

from openpyxl import Workbook, load_workbook

from wired_part.database.models import Part
from wired_part.database.repository import Repository
from wired_part.io.csv_handler import PART_CSV_COLUMNS, JOB_CSV_COLUMNS
from wired_part.io.validators import validate_part_row


def export_parts_excel(repo: Repository, filepath: str | Path) -> int:
    """Export all parts to an Excel workbook. Returns row count."""
    parts = repo.get_all_parts()
    filepath = Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Parts"

    # Header
    headers = [
        "Part #", "Description", "Quantity", "Min Qty",
        "Location", "Category", "Unit Cost", "Supplier", "Notes",
    ]
    ws.append(headers)

    for part in parts:
        ws.append([
            part.part_number,
            part.description,
            part.quantity,
            part.min_quantity,
            part.location,
            part.category_name,
            part.unit_cost,
            part.supplier,
            part.notes,
        ])

    # Auto-fit column widths (approximate)
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    wb.save(filepath)
    return len(parts)


def export_jobs_excel(repo: Repository, filepath: str | Path) -> int:
    """Export all jobs to an Excel workbook. Returns row count."""
    jobs = repo.get_all_jobs()
    filepath = Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    headers = ["Job #", "Name", "Customer", "Address", "Status", "Notes"]
    ws.append(headers)

    for job in jobs:
        ws.append([
            job.job_number,
            job.name,
            job.customer,
            job.address,
            job.status,
            job.notes,
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    wb.save(filepath)
    return len(jobs)


def import_parts_excel(
    repo: Repository,
    filepath: str | Path,
    update_existing: bool = False,
) -> dict:
    """Import parts from Excel. Returns results dict."""
    filepath = Path(filepath)
    results = {"imported": 0, "updated": 0, "skipped": 0, "errors": []}

    categories = {c.name: c.id for c in repo.get_all_categories()}

    try:
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            results["errors"].append("Empty workbook")
            return results

        # Use first row as header
        header = [str(h or "").strip().lower().replace(" ", "_").replace("#", "number")
                  for h in rows[0]]

        # Map common variations
        header_map = {
            "part_number": "part_number",
            "part_#": "part_number",
            "qty": "quantity",
            "min_qty": "min_quantity",
        }
        header = [header_map.get(h, h) for h in header]

        for row_num, row_data in enumerate(rows[1:], start=2):
            row = dict(zip(header, [str(v) if v is not None else "" for v in row_data]))

            errors = validate_part_row(row, row_num)
            if errors:
                results["errors"].extend(errors)
                results["skipped"] += 1
                continue

            pn = row.get("part_number", "").strip()
            existing = repo.get_part_by_number(pn)
            category_id = categories.get(row.get("category", "").strip())

            part = Part(
                id=existing.id if existing else None,
                part_number=pn,
                description=row.get("description", "").strip(),
                quantity=int(float(row.get("quantity", 0) or 0)),
                min_quantity=int(float(row.get("min_quantity", 0) or 0)),
                location=row.get("location", "").strip(),
                category_id=category_id,
                unit_cost=float(row.get("unit_cost", 0) or 0),
                supplier=row.get("supplier", "").strip(),
                notes=row.get("notes", "").strip(),
            )

            if existing and update_existing:
                repo.update_part(part)
                results["updated"] += 1
            elif existing:
                results["skipped"] += 1
            else:
                repo.create_part(part)
                results["imported"] += 1

        wb.close()

    except Exception as e:
        results["errors"].append(f"File error: {e}")

    return results
